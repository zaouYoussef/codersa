# -*- coding: utf-8 -*-
"""
Importe les feuilles JJ-MM-2026 de explication.xlsx vers
EXCEL_SAISIE_DATA et PRODUCTION['MAI'] dans Tableau_de_bord_Production.html.

· 01-05-2026 : cumuls d'ouverture r0_AC / r1_AC (AC − AB Excel).
· Autres jours : heures + champs saisissables seulement (totaux/cumuls calculés par le site).
"""
from __future__ import annotations

import json
import os
import re
import sys

import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
HTML_PATH = os.path.join(BASE, "Tableau_de_bord_Production.html")
XLSX_PATH = os.path.join(BASE, "explication.xlsx")

COL_B, COL_D, COL_F, COL_G, COL_H, COL_I, COL_J, COL_K, COL_L = 2, 4, 6, 7, 8, 9, 10, 11, 12
COL_M, COL_W = 13, 23
COL_Y, COL_Z, COL_AB, COL_AC = 25, 26, 28, 29
N_HOURS = 11
FIRST_DAY_MAI = "01-05-2026"
MONTH_KEY = "MAI"


def _num(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v) if not isinstance(v, bool) else None
    return None


def _hours_row(ws, row: int) -> list:
    out = []
    for c in range(COL_M, COL_W + 1):
        out.append(_num(ws.cell(row, c).value))
    while len(out) < N_HOURS:
        out.append(None)
    return out[:N_HOURS]


def _str_or_none(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def parse_daily_sheet(ws) -> list[dict]:
    blocks = []
    max_row = ws.max_row or 200
    for r in range(6, max_row + 1):
        lab = ws.cell(r, COL_L).value
        if lab != "Sortie Chaine":
            continue

        entre_r = r - 1
        eff_r, stag_r, form_r = r + 1, r + 2, r + 3
        if ws.cell(eff_r, COL_L).value != "Effectifs":
            eff_r = stag_r = form_r = None
        else:
            if ws.cell(stag_r, COL_L).value != "Stagiaire":
                stag_r = form_r = None
            elif form_r <= max_row and ws.cell(form_r, COL_L).value != "Formation":
                form_r = None

        chain = _num(ws.cell(r, COL_B).value)
        if chain is None:
            chain = _num(ws.cell(entre_r, COL_B).value)
        if chain is None:
            continue

        modele = _str_or_none(ws.cell(entre_r, COL_G).value) or _str_or_none(ws.cell(r, COL_G).value)
        import_cod = _str_or_none(ws.cell(r, COL_D).value)
        if not modele and import_cod:
            modele = import_cod
        if not modele:
            continue

        client = _str_or_none(ws.cell(r, COL_F).value) or ""
        qte = _num(ws.cell(r, COL_H).value)
        desig = _str_or_none(ws.cell(r, COL_I).value) or ""
        vt = _num(ws.cell(r, COL_J).value)
        ph = _num(ws.cell(r, COL_K).value)

        col_ab, col_ac = 28, 29
        ab_entre = _num(ws.cell(row=entre_r, column=col_ab).value) or 0
        ac_entre = _num(ws.cell(row=entre_r, column=col_ac).value)
        ab_sortie = _num(ws.cell(row=r, column=col_ab).value) or 0
        ac_sortie = _num(ws.cell(row=r, column=col_ac).value)

        r2_size = None
        if eff_r:
            h_eff = ws.cell(eff_r, COL_H).value
            if h_eff is not None and str(h_eff).strip().upper() in ("L", "S"):
                r2_size = str(h_eff).strip().upper()

        chef = _str_or_none(ws.cell(stag_r, COL_D).value) if stag_r else None
        k_obj = _num(ws.cell(eff_r, COL_K).value) if eff_r else None

        # 2em° = colonne Y uniquement (ligne Sortie) — vide sur toutes les feuilles terrain
        deux_sortie = _num(ws.cell(r, COL_Y).value)
        ae_2em = _num(ws.cell(row=r, column=31).value)
        # Colonne Z ligne Stagiaire = terme H.Lancement (AA), pas « 2em° »
        hlanc_z = _num(ws.cell(stag_r, COL_Z).value) if stag_r else None
        # EFF équilibrage (bloc VT/P/H) : colonne K de la ligne Stagiaire (ex. 54), pas le total effectifs
        eff_equilibre = _num(ws.cell(stag_r, COL_K).value) if stag_r else None

        excel_block = {
            "chain": int(chain),
            "import": import_cod or "",
            "client": client,
            "modele": modele,
            "qte_prevu": qte,
            "desig": desig,
            "r0_hours": _hours_row(ws, entre_r),
            "r1_ph": ph,
            "r1_meth": vt,
            "r1_hours": _hours_row(ws, r),
            "r2_size": r2_size,
            "r2_hours": _hours_row(ws, eff_r) if eff_r else [None] * N_HOURS,
            "r3_chef": chef,
            "r3_K_obj": k_obj,
            "r3_L_lab": "Stagiaire",
            "r3_hours": _hours_row(ws, stag_r) if stag_r else [None] * N_HOURS,
            "r4_hours": _hours_row(ws, form_r) if form_r else [None] * N_HOURS,
        }
        if ac_entre is not None:
            excel_block["r0_AC"] = ac_entre - ab_entre
        if ac_sortie is not None:
            excel_block["r1_AC"] = ac_sortie - ab_sortie
        if eff_equilibre is not None:
            excel_block["r2_eff"] = eff_equilibre

        prod_row = {
            "chain": int(chain),
            "import": import_cod or "",
            "client": client,
            "client_norm": client.upper(),
            "modele": modele,
            "qte_prevu": qte,
            "desig": desig,
            "methodes": vt,
            "ph": ph,
            "chef": chef or "",
            "chef_norm": (chef or "").upper(),
            "hours_entry": excel_block["r0_hours"][:],
            "hours_exit": excel_block["r1_hours"][:],
            "effectifs_hours": excel_block["r2_hours"][:],
            "stagiaire_hours": excel_block["r3_hours"][:],
            "formation_hours": excel_block["r4_hours"][:],
        }
        if ae_2em is not None:
            excel_block["r1_cumul_2em_ae"] = ae_2em
        if deux_sortie is not None:
            prod_row["deuxieme"] = deux_sortie
            excel_block["deuxieme"] = deux_sortie
        if hlanc_z is not None:
            prod_row["hlanc_z"] = hlanc_z
            excel_block["hlanc_z"] = hlanc_z
        if r2_size:
            prod_row["r2_size"] = r2_size
        if eff_equilibre is not None:
            prod_row["effectifs"] = eff_equilibre

        blocks.append((excel_block, prod_row, ac_entre, ab_entre, ac_sortie, ab_sortie))

    return blocks


def main() -> int:
    if not os.path.isfile(XLSX_PATH):
        print("Manquant:", XLSX_PATH, file=sys.stderr)
        return 1
    if not os.path.isfile(HTML_PATH):
        print("Manquant:", HTML_PATH, file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    date_sheets = sorted(
        [s for s in wb.sheetnames if re.match(r"^\d{2}-\d{2}-2026$", s)],
        key=lambda s: (int(s[6:10]), int(s[3:5]), int(s[0:2])),
    )
    may_sheets = [s for s in date_sheets if s.endswith("-05-2026")]
    print("Feuilles mai:", may_sheets)

    excel_by_date: dict[str, list] = {}
    prod_by_date: dict[str, list] = {}

    for sn in may_sheets:
        ws = wb[sn]
        h_moy_sortie = _num(ws.cell(9, 40).value)
        h_hlanc_ref = _num(ws.cell(3, 46).value)
        parsed = parse_daily_sheet(ws)
        excel_blocks = []
        prod_rows = []
        for excel_b, prod_b, ac_e, ab_e, ac_s, ab_s in parsed:
            eb = dict(excel_b)
            if h_moy_sortie is not None:
                eb["h_moy_sortie"] = h_moy_sortie
            if h_hlanc_ref is not None:
                eb["h_hlanc_ref"] = h_hlanc_ref
            excel_blocks.append(eb)
            prod_rows.append(prod_b)
        excel_by_date[sn] = excel_blocks
        prod_by_date[sn] = prod_rows
        print(f"  {sn}: {len(excel_blocks)} blocs")

    with open(HTML_PATH, encoding="utf-8") as f:
        html = f.read()

    m_ex = re.search(r"(window\.EXCEL_SAISIE_DATA = )(\{[\s\S]*?\})(;\s*\n)", html)
    if not m_ex:
        print("EXCEL_SAISIE_DATA introuvable", file=sys.stderr)
        return 1
    excel_all = json.loads(m_ex.group(2))
    for sn in may_sheets:
        excel_all[sn] = excel_by_date[sn]
    excel_json = json.dumps(excel_all, ensure_ascii=False, separators=(",", ":"))
    html = html[: m_ex.start(2)] + excel_json + html[m_ex.end(2) :]

    m_pr = re.search(r"(const PRODUCTION = )(\{[\s\S]*?\})(;\s*const CHART_FONT)", html)
    if not m_pr:
        print("PRODUCTION introuvable", file=sys.stderr)
        return 1
    prod_all = json.loads(m_pr.group(2))
    if MONTH_KEY not in prod_all:
        prod_all[MONTH_KEY] = {"label": MONTH_KEY, "daily_sheets": [], "daily": {}}
    prod_all[MONTH_KEY]["daily_sheets"] = may_sheets
    prod_all[MONTH_KEY]["daily"] = prod_by_date
    prod_json = json.dumps(prod_all, ensure_ascii=False, separators=(",", ":"))
    html = html[: m_pr.start(2)] + prod_json + html[m_pr.end(2) :]

    with open(HTML_PATH, "w", encoding="utf-8") as f:
        f.write(html)

    print("OK —", HTML_PATH)
    print("MAI:", len(may_sheets), "jours,", sum(len(v) for v in prod_by_date.values()), "lignes")
    return 0


if __name__ == "__main__":
    sys.exit(main())
