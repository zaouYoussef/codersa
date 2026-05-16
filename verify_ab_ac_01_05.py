# -*- coding: utf-8 -*-
import json
import re
from pathlib import Path

import openpyxl

from import_explication_mai import parse_daily_sheet, apply_first_day_cumuls, COL_B, COL_L, _num

html = (Path(__file__).parent / "Tableau_de_bord_Production.html").read_text(encoding="utf-8")
ex = json.loads(re.search(r"window\.EXCEL_SAISIE_DATA = (\{[\s\S]*?\});\s*\n", html).group(1))["01-05-2026"]
wb = openpyxl.load_workbook(Path(__file__).parent / "explication.xlsx", data_only=True)
ws = wb["01-05-2026"]
errs = []
for excel_b, _, ac_e, ab_e, ac_s, ab_s in parse_daily_sheet(ws):
    eb = dict(excel_b)
    apply_first_day_cumuls(eb, ac_e, ab_e, ac_s, ab_s)
    hb = next((x for x in ex if x["chain"] == eb["chain"] and x["modele"] == eb["modele"]), None)
    if not hb:
        errs.append("missing " + eb["modele"])
        continue
    te = sum(v or 0 for v in eb["r0_hours"])
    ts = sum(v or 0 for v in eb["r1_hours"])
    if hb.get("r0_AC") != eb.get("r0_AC"):
        errs.append(f"Ch{eb['chain']:02d} r0_AC HTML={hb.get('r0_AC')} attendu={eb.get('r0_AC')}")
    if hb.get("r1_AC") != eb.get("r1_AC"):
        errs.append(f"Ch{eb['chain']:02d} r1_AC HTML={hb.get('r1_AC')} attendu={eb.get('r1_AC')}")
    for r in range(6, 250):
        if ws.cell(r, COL_L).value != "Sortie Chaine":
            continue
        ch = _num(ws.cell(r, COL_B).value) or _num(ws.cell(r - 1, COL_B).value)
        if ch is None or int(ch) != eb["chain"]:
            continue
        m = str(ws.cell(r - 1, 7).value or ws.cell(r, 7).value or "")
        if m != eb["modele"]:
            continue
        xab = ws.cell(r, 28).value
        xac = ws.cell(r, 29).value
        xeab = ws.cell(r - 1, 28).value
        xeac = ws.cell(r - 1, 29).value
        if abs((xab or 0) - ts) > 0.01:
            errs.append(f"Ch{eb['chain']:02d} {m[:14]} AB sortie excel={xab} site={ts}")
        if abs((xac or 0) - (hb.get("r1_AC", 0) + ts)) > 0.01:
            errs.append(f"Ch{eb['chain']:02d} AC sortie excel={xac} site={hb.get('r1_AC') + ts}")
        if abs((xeab or 0) - te) > 0.01:
            errs.append(f"Ch{eb['chain']:02d} AB entre excel={xeab} site={te}")
        if abs((xeac or 0) - (hb.get("r0_AC", 0) + te)) > 0.01:
            errs.append(f"Ch{eb['chain']:02d} AC entre excel={xeac} site={hb.get('r0_AC') + te}")
        break

print("lignes", len(ex))
print("erreurs AB/AC/cumuls", len(errs))
for e in errs:
    print(" ", e)
if not errs:
    print("OK - donnees brutes et cumuls 01-05 conformes a explication.xlsx")
