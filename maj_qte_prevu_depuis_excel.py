# -*- coding: utf-8 -*-
"""
Met à jour le bloc PRODUCTION dans Tableau_de_bord_Production.html :
  - qte_prevu lue sur les lignes « Sortie Chaine » des feuilles JJ-MM-AAAA (colonne Qté prévue).
  - à défaut de valeur sur la ligne jour : « Plan Entrée chaines » de la feuille Encours au Prép.

Lancer depuis le dossier codersa :
  python maj_qte_prevu_depuis_excel.py

Les classeurs attendus : *Prod*Chaine*MM-YYYY*.xlsx (ex. EXL009-20 Prod H_Chaine 01-2026 -.xlsx → JANVIER).
"""
from __future__ import annotations

import glob
import json
import os
import re
import sys

import openpyxl

MONTH_FR = {
    "01": "JANVIER",
    "02": "FEVRIER",
    "03": "MARS",
    "04": "AVRIL",
    "05": "MAI",
    "06": "JUIN",
    "07": "JUILLET",
    "08": "AOUT",
    "09": "SEPTEMBRE",
    "10": "OCTOBRE",
    "11": "NOVEMBRE",
    "12": "DECEMBRE",
}


def month_key_from_filename(name: str) -> tuple[str | None, str | None]:
    """Retourne (clé PRODUCTION, année 'YYYY') ou (None, None)."""
    m = re.search(r"Chaine\s+(\d{2})-(\d{4})", name, re.I)
    if not m:
        return None, None
    mo = MONTH_FR.get(m.group(1))
    return mo, m.group(2)


def _norm_modele(x) -> str:
    if x is None:
        return ""
    return str(x).strip()


def encours_plan_by_model(wb: openpyxl.Workbook) -> dict[str, float]:
    out: dict[str, float] = {}
    sheet_name = None
    for n in wb.sheetnames:
        if "Encours" in n and "Pr" in n:
            sheet_name = n
            break
    if not sheet_name:
        return out
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=6, max_row=800, values_only=True):
        r = list(row) + [None] * 8
        modele, plan = r[4], r[5]
        m = _norm_modele(modele)
        if not m or m.lower() == "total":
            continue
        if isinstance(plan, (int, float)) and plan > 0:
            out[m] = max(out.get(m, 0.0), float(plan))
    return out


def daily_max_qte_by_model(wb: openpyxl.Workbook, year: str) -> dict[str, float]:
    """Parcourt les feuilles JJ-MM-YYYY du même millésime."""
    out: dict[str, float] = {}
    pat = re.compile(rf"^\d{{2}}-\d{{2}}-{year}$")
    for sn in wb.sheetnames:
        if not pat.match(sn):
            continue
        ws = wb[sn]
        for row in ws.iter_rows(values_only=True):
            r = list(row) + [None] * 28
            si = None
            for j, v in enumerate(r[:28]):
                if v == "Sortie Chaine":
                    si = j
                    break
            if si is None or si < 5:
                continue
            modele = r[si - 5]
            qte = r[si - 4]
            m = _norm_modele(modele)
            if not m or len(m) < 2:
                continue
            if isinstance(qte, (int, float)) and qte > 0:
                out[m] = max(out.get(m, 0.0), float(qte))
    return out


def merge_sources(enc: dict[str, float], daily: dict[str, float]) -> dict[str, float]:
    merged = dict(enc)
    for k, v in daily.items():
        merged[k] = max(merged.get(k, 0.0), v)
    return merged


def apply_to_month(payload: dict, month_key: str, qte_by_model: dict[str, float]) -> int:
    n = 0
    block = payload.get(month_key)
    if not block or "daily" not in block:
        return 0
    for _date, rows in block["daily"].items():
        for row in rows:
            m = _norm_modele(row.get("modele"))
            if not m:
                continue
            q = qte_by_model.get(m)
            if not q or q <= 0:
                continue
            cur = row.get("qte_prevu")
            curf = float(cur) if isinstance(cur, (int, float)) and cur > 0 else 0.0
            row["qte_prevu"] = max(curf, q)
            n += 1
    return n


def main() -> int:
    base = os.path.dirname(os.path.abspath(__file__))
    html_path = os.path.join(base, "Tableau_de_bord_Production.html")
    if not os.path.isfile(html_path):
        print("Fichier HTML introuvable:", html_path, file=sys.stderr)
        return 1

    with open(html_path, encoding="utf-8") as f:
        html = f.read()

    m = re.search(r"(const PRODUCTION = )(\{[\s\S]*?\})(;\s*const CHART_FONT)", html)
    if not m:
        print("Bloc const PRODUCTION = ... introuvable.", file=sys.stderr)
        return 1

    payload = json.loads(m.group(2))

    xlsx_files = [
        p
        for p in glob.glob(os.path.join(base, "*.xlsx"))
        if "Chaine" in os.path.basename(p) and "Prod" in os.path.basename(p)
    ]
    if not xlsx_files:
        print("Aucun classeur *Prod*Chaine*.xlsx dans", base, file=sys.stderr)
        return 1

    total_updates = 0
    for xlsx in sorted(xlsx_files):
        mk, year = month_key_from_filename(os.path.basename(xlsx))
        if not mk or mk not in payload:
            print("Ignoré (mois non présent dans PRODUCTION):", os.path.basename(xlsx))
            continue
        wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
        enc = encours_plan_by_model(wb)
        daily = daily_max_qte_by_model(wb, year)
        wb.close()
        merged = merge_sources(enc, daily)
        n = apply_to_month(payload, mk, merged)
        total_updates += n
        print(
            f"{os.path.basename(xlsx)} -> {mk}: "
            f"Encours {len(enc)} modeles, jour max {len(daily)} modeles, "
            f"{len(merged)} apres fusion, {n} lignes daily mises a jour."
        )

    new_json = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    new_html = html[: m.start(2)] + new_json + html[m.end(2) :]
    with open(html_path, "w", encoding="utf-8", newline="\n") as f:
        f.write(new_html)

    print("OK:", html_path, "| lignes mises a jour (qte_prevu):", total_updates)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
