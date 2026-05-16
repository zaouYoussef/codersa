# -*- coding: utf-8 -*-
import json
import re
from pathlib import Path

html = (Path(__file__).parent / "Tableau_de_bord_Production.html").read_text(encoding="utf-8")
ex = json.loads(re.search(r"window\.EXCEL_SAISIE_DATA = (\{[\s\S]*?\});\s*\n", html).group(1))["01-05-2026"]
prod = json.loads(re.search(r"const PRODUCTION = (\{[\s\S]*?\})(;\s*const CHART_FONT)", html).group(1))["MAI"]["daily"]["01-05-2026"]

EXCEL_AN9 = 8.25
EXCEL_AQ3 = 8.75


def sum_arr(a):
    return sum(0 if v is None else float(v) for v in a)


def count_a(a):
    return sum(1 for v in a[:11] if v is not None)


def hours_mw(a):
    return (a or [])[:11]


def calc_total_mw_u(h):
    hrs = hours_mw(h)
    u = hrs[8] if len(hrs) > 8 else None
    return sum_arr(hrs) if u is None else sum_arr(hrs) - float(u) * 0.25


def find_block(ch, mod):
    mod = mod.upper()
    for b in ex:
        if b["chain"] == ch and mod in b["modele"].upper():
            return b
    return None


def chain_sortie(ch):
    s = 0
    for m in prod:
        if m["chain"] != ch:
            continue
        b = find_block(ch, m["modele"])
        if b:
            s += sum_arr(b["r1_hours"])
    return s


def moy_sortie(r2, tot, r1, sumif):
    hrs = hours_mw(r1)
    ca = count_a(hrs)
    cs = count_a(hrs[:9])
    if str(r2).lower() == "s":
        return 0
    if ca > 9 and cs == 9:
        return round(sum_arr(hrs[:9]) / 9)
    if ca == 9 and cs == 9 and tot > 0:
        return round(tot / EXCEL_AQ3)
    if tot == 0:
        return 0
    if ca == 9:
        return sumif / EXCEL_AN9
    return sumif / ca if ca else 0


def hlanc_z(line, tsr, te, ph):
    if line.get("deuxieme_stag") is not None:
        return float(line["deuxieme_stag"])
    if tsr > 0:
        return (tsr / 4) * (EXCEL_AQ3 / EXCEL_AN9)
    k = float(ph or 0)
    if te > 0 and k > 0:
        return (te / k) * (EXCEL_AQ3 / 2)
    return 0


def hlanc(r1, tot, ph, line, tsr, te):
    c = count_a(hours_mw(r1))
    k = float(ph)
    ab = float(tot)
    z = hlanc_z(line, tsr, te, ph)
    if not k or not ab:
        return z
    r = ab / k
    core = EXCEL_AQ3 - r if c == 9 else (c - r if c > 0 else 0)
    return core + z


spots = [
    (1, "2843-709", {"moy": 55, "hlanc": 18.62, "tot_s": 479, "cum_s": 7024, "enc": 252, "rend": 27}),
    (2, "7696-837", {"moy": 63, "hlanc": 32.61, "tot_s": 551, "cum_s": 6136, "enc": 486}),
    (5, "2843-709", {"moy": 60, "hlanc": 36.12, "tot_s": 522, "cum_s": 7042}),
    (14, "TESS KEYHOLE", {"moy": 80, "tot_s": 343, "cum_s": 2757}),
    (15, "2878-556", {"moy": 75, "hlanc": 32.66, "rend": 224}),
]

print("=== Comparaison site (formules) vs capture utilisateur ===\n")
for ch, mod, exp in spots:
    b = find_block(ch, mod)
    if not b:
        print(f"Ch{ch:02d} {mod}: INTROUVABLE")
        continue
    r1 = b["r1_hours"]
    te = sum_arr(b["r0_hours"])
    ts = sum_arr(r1)
    ae = b.get("r0_AC", 0)
    ac = b.get("r1_AC", 0)
    csum = chain_sortie(ch)
    ms = moy_sortie(b.get("r2_size", ""), ts, r1, csum)
    teff = calc_total_mw_u(b["r2_hours"])
    tstag = calc_total_mw_u(b["r3_hours"])
    tsraw = sum_arr(hours_mw(b["r3_hours"]))
    h = hlanc(r1, ts, b["r1_ph"], b, tsraw, teff)
    rend = (b["r1_meth"] * 110) / (teff + tstag / 2) / 100 if teff + tstag / 2 else None
    name = b["modele"][:18]
    print(f"Ch{ch:02d} {name}")
    for k, ui in exp.items():
        got = {"moy": ms, "hlanc": round(h, 2), "tot_s": ts, "cum_s": ac + ts, "enc": ae + te - (ac + ts), "rend": round(rend * 100) if rend else None}[k]
        ok = "OK" if (got == ui or (isinstance(got, float) and abs(got - ui) < 0.05)) else "ECART"
        print(f"  {k:8} site={got}  UI={ui}  {ok}")

# Excel AB/AC raw
import openpyxl
from import_explication_mai import COL_L, _num

wb = openpyxl.load_workbook(Path(__file__).parent / "explication.xlsx", data_only=True)
ws = wb["01-05-2026"]
print("\n=== AB/AC Excel vs site (Ch01 2843) ===")
for r in range(6, 200):
    if ws.cell(r, COL_L).value != "Sortie Chaine":
        continue
    modele = str(ws.cell(r - 1, 7).value or "")
    if "2843-709" not in modele:
        continue
    b = find_block(1, "2843")
    print(f"  Excel AB sortie={ws.cell(r,28).value} AC={ws.cell(r,29).value}")
    print(f"  Site  AB sortie={sum_arr(b['r1_hours'])} cumul={b['r1_AC']+sum_arr(b['r1_hours'])}")
    print(f"  Excel X moy={ws.cell(r,24).value}  Site moy={moy_sortie(b.get('r2_size',''), sum_arr(b['r1_hours']), b['r1_hours'], chain_sortie(1))}")
    break
