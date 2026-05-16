# -*- coding: utf-8 -*-
"""Vérifie 01-05-2026 : explication.xlsx vs HTML + formules site."""
from __future__ import annotations

import json
import math
import re
import sys
from pathlib import Path

import openpyxl

from import_explication_mai import (
    FIRST_DAY_MAI,
    parse_daily_sheet,
    apply_first_day_cumuls,
    _num,
    COL_B,
    COL_L,
)

BASE = Path(__file__).resolve().parent
HTML = BASE / "Tableau_de_bord_Production.html"
XLSX = BASE / "explication.xlsx"
DATE = FIRST_DAY_MAI

EXCEL_AN9 = 8.25
EXCEL_AQ3 = 8.75
EXCEL_MOY_JOUR_PLEIN = EXCEL_AN9 + 0.75
EXCEL_HLANC_COUNT_FULL = EXCEL_AQ3 + 0.25
EXCEL_COL_U_INDEX = 8
EXCEL_U_DEDUCT_MULT = 0.25
EXCEL_EFF_COUNT_OFFSET = 0.25


def sum_arr(arr):
    return sum(0 if v is None or (isinstance(v, float) and math.isnan(v)) else float(v) for v in arr)


def count_a(arr):
    return sum(1 for v in arr[:11] if v is not None and v != "")


def hours_mw(arr):
    return (arr or [])[:11]


def calc_total_row_mw_u(hours):
    hrs = hours_mw(hours)
    u = hrs[EXCEL_COL_U_INDEX] if len(hrs) > EXCEL_COL_U_INDEX else None
    if u is None or u == "":
        return sum_arr(hrs)
    return sum_arr(hrs) - float(u) * EXCEL_U_DEDUCT_MULT


def calc_moy_sortie_display(r2_size, total_sortie, r1h, sumif_sortie):
    hrs = hours_mw(r1h)
    std = hrs[: int(EXCEL_MOY_JOUR_PLEIN)]
    count_all = count_a(hrs)
    count_std = count_a(std)
    if str(r2_size).lower() == "s":
        return 0
    total = float(total_sortie or 0)
    if count_all > EXCEL_MOY_JOUR_PLEIN and count_std == EXCEL_MOY_JOUR_PLEIN:
        return round(sum_arr(std) / EXCEL_MOY_JOUR_PLEIN)
    if count_all == EXCEL_MOY_JOUR_PLEIN and count_std == EXCEL_MOY_JOUR_PLEIN and total > 0:
        return round(total / EXCEL_AQ3)
    sumif = sumif_sortie if isinstance(sumif_sortie, (int, float)) else total
    if total == 0:
        return 0
    if count_all == EXCEL_MOY_JOUR_PLEIN:
        return sumif / EXCEL_AN9
    if count_all > 0:
        return sumif / count_all
    return 0


def calc_hlanc_z(line, total_stag_raw, total_eff, ph):
    if line.get("deuxieme_stag") is not None:
        return float(line["deuxieme_stag"])
    if total_stag_raw > 0:
        return (total_stag_raw / 4) * (EXCEL_AQ3 / EXCEL_AN9)
    k = float(ph or 0)
    if total_eff > 0 and k > 0:
        return (total_eff / k) * (EXCEL_AQ3 / 2)
    return 0


def calc_hlanc(r1h, total_sortie, ph, line, total_stag_raw, total_eff):
    hrs = hours_mw(r1h)
    count = count_a(hrs)
    k = float(ph or 0)
    ab = float(total_sortie or 0)
    z = calc_hlanc_z(line, total_stag_raw, total_eff, ph)
    if not k or not ab:
        return z if math.isfinite(z) else None
    ratio = ab / k
    if count == EXCEL_HLANC_COUNT_FULL:
        core = EXCEL_AQ3 - ratio
    elif count > 0:
        core = count - ratio
    else:
        core = 0
    return core + z


def compute_block(line, excel, chain_total_sortie=None):
    r0h = line.get("hours_entry") or excel.get("r0_hours") or []
    r1h = line.get("hours_exit") or excel.get("r1_hours") or []
    r2h = line.get("effectifs_hours") or excel.get("r2_hours") or []
    r3h = line.get("stagiaire_hours") or excel.get("r3_hours") or []
    total_entre = sum_arr(r0h)
    total_sortie = sum_arr(r1h)
    total_eff = calc_total_row_mw_u(r2h)
    total_stag_raw = sum_arr(hours_mw(r3h))
    total_stag = calc_total_row_mw_u(r3h)
    ac_entre = float(excel.get("r0_AC") or 0)
    ac_sortie = float(excel.get("r1_AC") or 0)
    cumul_entre = ac_entre + total_entre
    cumul_sortie = ac_sortie + total_sortie
    encours = cumul_entre - cumul_sortie
    vt = line.get("methodes") or excel.get("r1_meth") or 0
    ph = line.get("ph") or excel.get("r1_ph") or 0
    denom = total_eff + total_stag / 2
    rend = (vt * 110) / denom / 100 if denom and vt else None
    r2_size = line.get("r2_size") or excel.get("r2_size") or ""
    sumif = chain_total_sortie if chain_total_sortie is not None else total_sortie
    moy_sortie = calc_moy_sortie_display(r2_size, total_sortie, r1h, sumif)
    hlanc = calc_hlanc(r1h, total_sortie, ph, {**excel, **line}, total_stag_raw, total_eff)
    qp = line.get("qte_prevu") or excel.get("qte_prevu")
    qte_reste = (float(qp) - cumul_entre) if qp is not None else None
    return {
        "total_entre": total_entre,
        "total_sortie": total_sortie,
        "total_eff": total_eff,
        "total_stag": total_stag,
        "cumul_entre": cumul_entre,
        "cumul_sortie": cumul_sortie,
        "encours": encours,
        "rend": rend,
        "moy_sortie": moy_sortie,
        "hlanc": hlanc,
        "qte_reste": qte_reste,
    }


def find_block(arr, chain, modele):
    mod = modele.strip().upper()
    for b in arr:
        if int(b["chain"]) == chain and b["modele"].strip().upper() == mod:
            return b
    return None


def read_excel_row_metrics(ws, sortie_r):
    entre_r = sortie_r - 1
    eff_r = sortie_r + 1
    stag_r = sortie_r + 2
    col_x, col_aa, col_ab, col_ac = 24, 27, 28, 29

    def cell(r, c):
        return _num(ws.cell(r, c).value)

    return {
        "entre_ab": cell(entre_r, col_ab),
        "entre_ac": cell(entre_r, col_ac),
        "sortie_ab": cell(sortie_r, col_ab),
        "sortie_ac": cell(sortie_r, col_ac),
        "moy_sortie_x": cell(sortie_r, col_x),
        "hlanc_aa": cell(stag_r, col_aa),
    }


def lists_eq(a, b):
    if not isinstance(a, list) or not isinstance(b, list):
        return a == b
    if len(a) != len(b):
        return False
    for va, vb in zip(a, b):
        if va is None and vb is None:
            continue
        if va is None or vb is None:
            return False
        if abs(float(va) - float(vb)) > 1e-6:
            return False
    return True


def near(a, b, tol=0.5):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    return abs(float(a) - float(b)) <= tol


def main():
    html = HTML.read_text(encoding="utf-8")
    excel_all = json.loads(re.search(r"window\.EXCEL_SAISIE_DATA = (\{[\s\S]*?\});\s*\n", html).group(1))
    prod_all = json.loads(re.search(r"const PRODUCTION = (\{[\s\S]*?\})(;\s*const CHART_FONT)", html).group(1))
    html_excel = excel_all[DATE]
    html_prod = prod_all["MAI"]["daily"][DATE]

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb[DATE]
    parsed = parse_daily_sheet(ws)

    errors = []
    warnings = []

    chain_sortie = {}
    for excel_b, _, _, _, _, _ in parsed:
        ch = excel_b["chain"]
        chain_sortie[ch] = chain_sortie.get(ch, 0) + sum_arr(excel_b["r1_hours"])

    for excel_b, prod_b, ac_e, ab_e, ac_s, ab_s in parsed:
        eb = dict(excel_b)
        apply_first_day_cumuls(eb, ac_e, ab_e, ac_s, ab_s)
        hb = find_block(html_excel, eb["chain"], eb["modele"])
        if not hb:
            errors.append(f"Manquant HTML: Ch{eb['chain']:02d} {eb['modele']}")
            continue
        for key in ("r0_hours", "r1_hours", "r2_hours", "r3_hours", "r1_ph", "r1_meth", "qte_prevu", "r0_AC", "r1_AC"):
            if key in ("r0_AC", "r1_AC"):
                if hb.get(key) != eb.get(key):
                    errors.append(f"Ch{eb['chain']:02d} {eb['modele'][:14]} {key}: HTML={hb.get(key)} xlsx={eb.get(key)}")
            elif not lists_eq(hb.get(key), eb.get(key)) if key.endswith("_hours") else (
                hb.get(key) != eb.get(key)
                and not (hb.get(key) is not None and eb.get(key) is not None and abs(float(hb.get(key)) - float(eb.get(key))) < 1e-6)
            ):
                errors.append(f"Ch{eb['chain']:02d} {eb['modele'][:14]} {key} diffère")

    for r in range(6, (ws.max_row or 200) + 1):
        if ws.cell(r, COL_L).value != "Sortie Chaine":
            continue
        entre_r = r - 1
        ch = _num(ws.cell(r, COL_B).value) or _num(ws.cell(entre_r, COL_B).value)
        if ch is None:
            continue
        chain = int(ch)
        modele = str(ws.cell(entre_r, 7).value or ws.cell(r, 7).value or "").strip()
        hb = find_block(html_excel, chain, modele)
        if not hb:
            continue
        hp = find_block(html_prod, chain, modele)
        xm = read_excel_row_metrics(ws, r)
        csum = chain_sortie.get(chain)
        calc = compute_block(hp or {}, hb, csum)
        label = f"Ch{chain:02d} {modele[:18]}"
        for name, got, exp, tol in [
            ("AB entre", calc["total_entre"], xm["entre_ab"], 0.5),
            ("AB sortie", calc["total_sortie"], xm["sortie_ab"], 0.5),
            ("AC entrée", calc["cumul_entre"], xm["entre_ac"], 1),
            ("AC sortie", calc["cumul_sortie"], xm["sortie_ac"], 1),
            ("Moy sortie", calc["moy_sortie"], xm["moy_sortie_x"], 0.15),
            ("H.Lanc", calc["hlanc"], xm["hlanc_aa"], 0.15),
        ]:
            if exp is None and (got is None or got == 0):
                continue
            if not near(got, exp, tol):
                errors.append(f"{label} {name}: site={got} excel={exp}")

    day_entre = day_sortie = day_eff = day_stag = 0
    rend_sum = rend_n = 0
    for p in html_prod:
        hb = find_block(html_excel, p["chain"], p["modele"])
        if not hb:
            continue
        c = compute_block(p, hb, chain_sortie.get(p["chain"]))
        day_entre += c["total_entre"]
        day_sortie += c["total_sortie"]
        day_eff += c["total_eff"]
        day_stag += c["total_stag"]
        if c["rend"]:
            rend_sum += c["rend"]
            rend_n += 1

    print(f"=== {DATE} ===")
    print(f"Lignes: {len(html_prod)} (xlsx: {len(parsed)})")
    print(f"Totaux site — Entrée: {day_entre:.0f}  Sortie: {day_sortie:.0f}  Eff: {day_eff:.2f}  Stag: {day_stag:.2f}")
    print(f"UI utilisateur — Entrée: 7399  Sortie: 7533  Eff: 6550  Stag: 19  Rdt moy: 25.6%")
    if abs(day_entre - 7399) > 1:
        warnings.append(f"Δ entrée {day_entre - 7399:+.0f}")
    if abs(day_sortie - 7533) > 1:
        warnings.append(f"Δ sortie {day_sortie - 7533:+.0f}")
    if abs(day_eff - 6550) > 2:
        warnings.append(f"Δ effectifs {day_eff - 6550:+.1f}")
    if abs(day_stag - 19) > 0.5:
        warnings.append(f"Δ stagiaires {day_stag - 19:+.1f}")

    ch01 = find_block(html_excel, 1, "2843-709-800")
    if ch01:
        c = compute_block({}, ch01, chain_sortie[1])
        print(f"\nRéf. Ch01 2843-709-800:")
        print(f"  cumuls ouverture r0_AC={ch01.get('r0_AC')} r1_AC={ch01.get('r1_AC')}")
        print(f"  jour AB entrée=0 sortie={c['total_sortie']:.0f} | cumul entrée={c['cumul_entre']:.0f} sortie={c['cumul_sortie']:.0f} encours={c['encours']:.0f}")
        print(f"  moy={c['moy_sortie']} hlanc={c['hlanc']:.2f} rend={c['rend']*100:.1f}% qte_reste={c['qte_reste']:.0f}")

    print(f"\nErreurs ({len(errors)}):")
    for e in errors[:50]:
        print(" ", e)
    if len(errors) > 50:
        print(f"  ... +{len(errors)-50}")
    if warnings:
        print(f"\nAvertissements:")
        for w in warnings:
            print(" ", w)
    return 1 if errors else 0


if __name__ == "__main__":
    sys.exit(main())
